"""
Router para operaciones CRUD de Control de Consumo Diario (Matriz 4)
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.database import get_db
from models.control_consumo_diario import ControlConsumoDiario
from schemas.control_consumo_diario import (
    ControlConsumoDiarioCreate,
    ControlConsumoDiarioUpdate,
    ControlConsumoDiarioResponse
)

router = APIRouter()


@router.get("/", response_model=List[ControlConsumoDiarioResponse])
def get_consumos_diarios(
    skip: int = 0,
    limit: int = 100,
    fecha: Optional[date] = None,
    quimico_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Obtener lista de consumos diarios"""
    query = db.query(ControlConsumoDiario)
    
    if fecha:
        query = query.filter(ControlConsumoDiario.fecha == fecha)
    
    if quimico_id:
        query = query.filter(ControlConsumoDiario.quimico_id == quimico_id)
    
    consumos = query.order_by(
        ControlConsumoDiario.fecha.desc()
    ).offset(skip).limit(limit).all()
    
    return consumos


@router.get("/fecha/{fecha_consulta}", response_model=List[ControlConsumoDiarioResponse])
def get_consumos_por_fecha(fecha_consulta: date, db: Session = Depends(get_db)):
    """Obtener todos los consumos de un día específico"""
    consumos = db.query(ControlConsumoDiario).filter(
        ControlConsumoDiario.fecha == fecha_consulta
    ).all()
    
    return consumos


@router.get("/{consumo_id}", response_model=ControlConsumoDiarioResponse)
def get_consumo_diario(consumo_id: int, db: Session = Depends(get_db)):
    """Obtener un consumo diario por ID"""
    consumo = db.query(ControlConsumoDiario).filter(
        ControlConsumoDiario.id == consumo_id
    ).first()
    
    if not consumo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Consumo diario con ID {consumo_id} no encontrado"
        )
    
    return consumo


@router.post("/", response_model=ControlConsumoDiarioResponse, status_code=status.HTTP_201_CREATED)
def create_consumo_diario(consumo: ControlConsumoDiarioCreate, db: Session = Depends(get_db)):
    """Crear un nuevo registro de consumo diario"""
    db_consumo = ControlConsumoDiario(**consumo.model_dump())
    db.add(db_consumo)
    db.commit()
    db.refresh(db_consumo)
    return db_consumo


@router.put("/{consumo_id}", response_model=ControlConsumoDiarioResponse)
def update_consumo_diario(
    consumo_id: int,
    consumo: ControlConsumoDiarioUpdate,
    db: Session = Depends(get_db)
):
    """Actualizar un consumo diario existente"""
    db_consumo = db.query(ControlConsumoDiario).filter(
        ControlConsumoDiario.id == consumo_id
    ).first()
    
    if not db_consumo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Consumo diario con ID {consumo_id} no encontrado"
        )
    
    update_data = consumo.model_dump(exclude_unset=True)
    
    for key, value in update_data.items():
        setattr(db_consumo, key, value)
    
    db.commit()
    db.refresh(db_consumo)
    return db_consumo


@router.delete("/{consumo_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_consumo_diario(consumo_id: int, db: Session = Depends(get_db)):
    """Eliminar un consumo diario"""
    db_consumo = db.query(ControlConsumoDiario).filter(
        ControlConsumoDiario.id == consumo_id
    ).first()
    
    if not db_consumo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Consumo diario con ID {consumo_id} no encontrado"
        )
    
    db.delete(db_consumo)
    db.commit()
    return None


@router.get("/exportar-excel/fecha/{fecha}")
def exportar_excel_consumo_diario(fecha: str, db: Session = Depends(get_db)):
    """Exportar consumo diario a Excel con formato específico"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Consumo Diario"
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Helper function para aplicar bordes y merge
    def apply_borders_and_merge(cell_range):
        start_cell, end_cell = cell_range.split(':')
        start_col = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = ''.join(filter(str.isalpha, end_cell))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        
        from openpyxl.utils import column_index_from_string
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col_idx, end_col_idx + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        ws.merge_cells(cell_range)
    
    # Título (Fila 1)
    ws['A1'] = 'MANCOMUNIDAD "LA ESPERANZA"'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    apply_borders_and_merge('A1:P1')
    
    # Subtítulo 1 (Fila 2)
    ws['A2'] = 'PLANTA DE TRATAMIENTO DE AGUA POTABLE "LA ESPERANZA"'
    ws['A2'].font = title_font
    ws['A2'].alignment = center_align
    apply_borders_and_merge('A2:P2')
    
    # Subtítulo 2 (Fila 3)
    ws['A3'] = 'CONTROL DIARIO DEL CONSUMO DE QUÍMICOS'
    ws['A3'].font = bold_font
    ws['A3'].alignment = center_align
    apply_borders_and_merge('A3:P3')
    
    # FECHA (Fila 4)
    ws['A4'] = 'FECHA:'
    ws['A4'].font = bold_font
    ws['A4'].alignment = left_align
    ws['A4'].border = thin_border
    
    # OPERADORES (Fila 5)
    ws['A5'] = 'OPERADORES:'
    ws['A5'].font = bold_font
    ws['A5'].alignment = left_align
    ws['A5'].border = thin_border
    
    # Parsear fecha y obtener registros
    try:
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de fecha inválido. Use YYYY-MM-DD"
        )
    
    # Llenar fecha y operador
    ws['B4'] = fecha_obj.strftime('%d-%m-%Y')
    ws['B4'].border = thin_border
    apply_borders_and_merge('B4:P4')
    
    # Obtener registros y operador
    registros = db.query(ControlConsumoDiario).filter(
        ControlConsumoDiario.fecha == fecha_obj
    ).all()
    
    operador_nombre = "N/A"
    if registros and registros[0].usuario:
        operador_nombre = registros[0].usuario.nombre_completo
    
    ws['B5'] = operador_nombre
    ws['B5'].border = thin_border
    apply_borders_and_merge('B5:P5')
    
    # Crear diccionario de registros por químico
    registros_dict = {}
    for registro in registros:
        if registro.quimico:
            registros_dict[registro.quimico.nombre] = registro
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col].width = 12
    
    current_row = 7
    
    # 1. SULFATO DE ALUMINIO
    current_row = crear_seccion_quimico(
        ws, current_row, "SULFATO DE ALUMINIO (Sacos de 25 Kg.)",
        registros_dict.get("SULFATO DE ALUMINIO"), True,
        bold_font, center_align, left_align, thin_border, apply_borders_and_merge
    )
    
    # 2. CAL
    current_row = crear_seccion_quimico(
        ws, current_row, "CAL (Sacos de 25 Kg.)",
        registros_dict.get("CAL"), False,
        bold_font, center_align, left_align, thin_border, apply_borders_and_merge
    )
    
    # 3. HIPOCLORITO DE CALCIO (solo bodega)
    current_row = crear_seccion_hipoclorito(
        ws, current_row, registros_dict.get("HIPOCLORITO DE CALCIO"),
        bold_font, center_align, left_align, thin_border, apply_borders_and_merge
    )
    
    # 4. FLOERGER
    current_row = crear_seccion_quimico(
        ws, current_row, "FLOERGER (Sacos de 25 Kg.)",
        registros_dict.get("FLOERGER"), False,
        bold_font, center_align, left_align, thin_border, apply_borders_and_merge
    )
    
    # OBSERVACIONES
    ws.cell(row=current_row, column=1).value = 'OBSERVACIONES:'
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = left_align
    apply_borders_and_merge(f'A{current_row}:P{current_row + 2}')
    ws.row_dimensions[current_row].height = 20
    ws.row_dimensions[current_row + 1].height = 20
    ws.row_dimensions[current_row + 2].height = 20
    
    # REVISADO
    rev_row = current_row + 4
    ws.cell(row=rev_row, column=1).value = 'REVISADO'
    ws.cell(row=rev_row, column=1).font = bold_font
    ws.cell(row=rev_row, column=1).alignment = center_align
    apply_borders_and_merge(f'A{rev_row}:P{rev_row}')
    ws.row_dimensions[rev_row].height = 25
    
    # Guardar en buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Nombre del archivo
    filename = f"consumo_diario_{fecha}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def crear_seccion_quimico(ws, start_row, titulo, registro, tiene_dos_tanques, 
                          bold_font, center_align, left_align, thin_border, apply_borders_and_merge):
    """Crear sección para SULFATO, CAL o FLOERGER"""
    current_row = start_row
    
    # Título del químico
    ws.cell(row=current_row, column=1).value = titulo
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = center_align
    apply_borders_and_merge(f'A{current_row}:P{current_row}')
    current_row += 1
    
    # BODEGA (izquierda) y TANQUE N°1 (derecha)
    # Headers BODEGA (columnas A-C)
    ws.cell(row=current_row, column=1).value = 'BODEGA'
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = center_align
    apply_borders_and_merge(f'A{current_row}:C{current_row}')
    
    # Headers TANQUE N°1
    if tiene_dos_tanques:
        ws.cell(row=current_row, column=4).value = 'TANQUE N° 1'
        ws.cell(row=current_row, column=4).font = bold_font
        ws.cell(row=current_row, column=4).alignment = center_align
        apply_borders_and_merge(f'D{current_row}:H{current_row}')
        
        # TANQUE N°2
        ws.cell(row=current_row, column=9).value = 'TANQUE N° 2'
        ws.cell(row=current_row, column=9).font = bold_font
        ws.cell(row=current_row, column=9).alignment = center_align
        apply_borders_and_merge(f'I{current_row}:M{current_row}')
    else:
        ws.cell(row=current_row, column=4).value = 'TANQUE N° 1'
        ws.cell(row=current_row, column=4).font = bold_font
        ws.cell(row=current_row, column=4).alignment = center_align
        apply_borders_and_merge(f'D{current_row}:P{current_row}')
    
    current_row += 1
    
    # Sub-headers
    # BODEGA sub-headers
    for col, label in enumerate(['INGRESA', 'EGRESA', 'STOCK'], start=1):
        ws.cell(row=current_row, column=col).value = label
        ws.cell(row=current_row, column=col).font = bold_font
        ws.cell(row=current_row, column=col).alignment = center_align
        ws.cell(row=current_row, column=col).border = thin_border
    
    # TANQUE N°1 sub-headers
    tanque_headers = ['HORA', 'LECTURA INICIAL', 'LECTURA FINAL', 'CONSUMO']
    if tiene_dos_tanques:
        for col, label in enumerate(tanque_headers, start=4):
            ws.cell(row=current_row, column=col).value = label
            ws.cell(row=current_row, column=col).font = bold_font
            ws.cell(row=current_row, column=col).alignment = center_align
            ws.cell(row=current_row, column=col).border = thin_border
        
        # TANQUE N°2 sub-headers
        for col, label in enumerate(tanque_headers, start=9):
            ws.cell(row=current_row, column=col).value = label
            ws.cell(row=current_row, column=col).font = bold_font
            ws.cell(row=current_row, column=col).alignment = center_align
            ws.cell(row=current_row, column=col).border = thin_border
    else:
        for col, label in enumerate(tanque_headers, start=4):
            ws.cell(row=current_row, column=col).value = label
            ws.cell(row=current_row, column=col).font = bold_font
            ws.cell(row=current_row, column=col).alignment = center_align
            ws.cell(row=current_row, column=col).border = thin_border
    
    current_row += 1
    
    # Datos (3-4 filas vacías)
    for _ in range(4):
        # BODEGA datos
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).alignment = center_align
        
        # TANQUE datos
        if tiene_dos_tanques:
            for col in range(4, 13):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = center_align
        else:
            for col in range(4, 8):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = center_align
        
        current_row += 1
    
    # Llenar primera fila con datos si existen
    if registro:
        data_row = current_row - 4
        ws.cell(row=data_row, column=1).value = registro.bodega_ingresa if registro.bodega_ingresa else ''
        ws.cell(row=data_row, column=2).value = registro.bodega_egresa if registro.bodega_egresa else ''
        ws.cell(row=data_row, column=3).value = registro.bodega_stock if registro.bodega_stock else ''
        
        if registro.tanque1_hora:
            ws.cell(row=data_row, column=4).value = registro.tanque1_hora.strftime('%H:%M') if registro.tanque1_hora else ''
        ws.cell(row=data_row, column=5).value = float(registro.tanque1_lectura_inicial) if registro.tanque1_lectura_inicial else ''
        ws.cell(row=data_row, column=6).value = float(registro.tanque1_lectura_final) if registro.tanque1_lectura_final else ''
        ws.cell(row=data_row, column=7).value = float(registro.tanque1_consumo) if registro.tanque1_consumo else ''
        
        if tiene_dos_tanques and registro.tanque2_hora:
            ws.cell(row=data_row, column=9).value = registro.tanque2_hora.strftime('%H:%M') if registro.tanque2_hora else ''
            ws.cell(row=data_row, column=10).value = float(registro.tanque2_lectura_inicial) if registro.tanque2_lectura_inicial else ''
            ws.cell(row=data_row, column=11).value = float(registro.tanque2_lectura_final) if registro.tanque2_lectura_final else ''
            ws.cell(row=data_row, column=12).value = float(registro.tanque2_consumo) if registro.tanque2_consumo else ''
    
    # Fila TOTAL
    ws.cell(row=current_row, column=1).value = 'TOTAL'
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = center_align
    
    if tiene_dos_tanques:
        apply_borders_and_merge(f'A{current_row}:H{current_row}')
        ws.cell(row=current_row, column=9).border = thin_border
        ws.cell(row=current_row, column=9).alignment = center_align
    else:
        apply_borders_and_merge(f'A{current_row}:G{current_row}')
        ws.cell(row=current_row, column=8).border = thin_border
        ws.cell(row=current_row, column=8).alignment = center_align
    
    current_row += 2
    return current_row


def crear_seccion_hipoclorito(ws, start_row, registro, bold_font, center_align, left_align, thin_border, apply_borders_and_merge):
    """Crear sección para HIPOCLORITO DE CALCIO (solo bodega)"""
    current_row = start_row
    
    # Título
    ws.cell(row=current_row, column=1).value = 'HIPOCLORITO DE CALCIO ( 45 Kg)'
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = center_align
    apply_borders_and_merge(f'A{current_row}:P{current_row}')
    current_row += 1
    
    # Header BODEGA
    ws.cell(row=current_row, column=1).value = 'BODEGA'
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = center_align
    apply_borders_and_merge(f'A{current_row}:D{current_row}')
    current_row += 1
    
    # Sub-headers
    for col, label in enumerate(['INGRESA', 'EGRESA', 'CONSUMO', 'STOCK'], start=1):
        ws.cell(row=current_row, column=col).value = label
        ws.cell(row=current_row, column=col).font = bold_font
        ws.cell(row=current_row, column=col).alignment = center_align
        ws.cell(row=current_row, column=col).border = thin_border
    current_row += 1
    
    # Datos (3 filas)
    for i in range(3):
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).alignment = center_align
        current_row += 1
    
    # Llenar primera fila con datos si existen
    if registro:
        data_row = current_row - 3
        ws.cell(row=data_row, column=1).value = registro.bodega_ingresa if registro.bodega_ingresa else ''
        ws.cell(row=data_row, column=2).value = registro.bodega_egresa if registro.bodega_egresa else ''
        ws.cell(row=data_row, column=3).value = float(registro.total_consumo) if registro.total_consumo else ''
        ws.cell(row=data_row, column=4).value = registro.bodega_stock if registro.bodega_stock else ''
    
    # Fila TOTAL
    ws.cell(row=current_row, column=1).value = 'TOTAL (45Kg.)'
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = center_align
    apply_borders_and_merge(f'A{current_row}:C{current_row}')
    ws.cell(row=current_row, column=4).border = thin_border
    ws.cell(row=current_row, column=4).alignment = center_align
    
    current_row += 2
    return current_row
