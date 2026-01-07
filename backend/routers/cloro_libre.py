"""
Router para operaciones CRUD de Control de Cloro Libre (Matriz 5)
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
from decimal import Decimal
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.database import get_db
from models.control_cloro_libre import ControlCloroLibre
from schemas.control_cloro_libre import (
    ControlCloroLibreCreate,
    ControlCloroLibreUpdate,
    ControlCloroLibreResponse
)

router = APIRouter()


@router.get("/", response_model=List[ControlCloroLibreResponse])
def get_controles_cloro(
    skip: int = 0,
    limit: int = 100,
    codigo: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Obtener lista de controles de cloro libre"""
    query = db.query(ControlCloroLibre)
    
    if codigo:
        query = query.filter(ControlCloroLibre.codigo == codigo)
    
    controles = query.order_by(
        ControlCloroLibre.fecha_mes.desc()
    ).offset(skip).limit(limit).all()
    
    return controles


@router.get("/saldo-actual", response_model=ControlCloroLibreResponse)
def get_saldo_actual(db: Session = Depends(get_db)):
    """Obtener el saldo actual más reciente de cloro libre"""
    control = db.query(ControlCloroLibre).order_by(
        ControlCloroLibre.fecha_mes.desc()
    ).first()
    
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No hay registros de cloro libre"
        )
    
    return control


@router.get("/{control_id}", response_model=ControlCloroLibreResponse)
def get_control_cloro(control_id: int, db: Session = Depends(get_db)):
    """Obtener un control de cloro libre por ID"""
    control = db.query(ControlCloroLibre).filter(
        ControlCloroLibre.id == control_id
    ).first()
    
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Control de cloro libre con ID {control_id} no encontrado"
        )
    
    return control


@router.post("/", response_model=ControlCloroLibreResponse, status_code=status.HTTP_201_CREATED)
def create_control_cloro(control: ControlCloroLibreCreate, db: Session = Depends(get_db)):
    """Crear un nuevo registro de control de cloro libre"""
    db_control = ControlCloroLibre(**control.model_dump())
    db.add(db_control)
    db.commit()
    db.refresh(db_control)
    return db_control


@router.put("/{control_id}", response_model=ControlCloroLibreResponse)
def update_control_cloro(
    control_id: int,
    control: ControlCloroLibreUpdate,
    db: Session = Depends(get_db)
):
    """Actualizar un control de cloro libre existente"""
    db_control = db.query(ControlCloroLibre).filter(
        ControlCloroLibre.id == control_id
    ).first()
    
    if not db_control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Control de cloro libre con ID {control_id} no encontrado"
        )
    
    update_data = control.model_dump(exclude_unset=True)
    
    for key, value in update_data.items():
        setattr(db_control, key, value)
    
    db.commit()
    db.refresh(db_control)
    return db_control


@router.delete("/{control_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_control_cloro(control_id: int, db: Session = Depends(get_db)):
    """Eliminar un control de cloro libre"""
    db_control = db.query(ControlCloroLibre).filter(
        ControlCloroLibre.id == control_id
    ).first()
    
    if not db_control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Control de cloro libre con ID {control_id} no encontrado"
        )
    
    db.delete(db_control)
    db.commit()
    return None


@router.get("/exportar-excel/mes/{fecha_mes}")
def exportar_excel(fecha_mes: str, db: Session = Depends(get_db)):
    """Exportar registros de cloro libre a Excel con formato específico"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro Reactivos"
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    bold_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Colores para encabezados
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    entra_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')  # Rojo/naranja
    sale_fill = PatternFill(start_color='4ECDC4', end_color='4ECDC4', fill_type='solid')  # Azul
    saldo_fill = PatternFill(start_color='FFE66D', end_color='FFE66D', fill_type='solid')  # Amarillo
    total_fill = PatternFill(start_color='95E1D3', end_color='95E1D3', fill_type='solid')  # Verde
    
    # Helper function para aplicar bordes y merge
    def apply_borders_and_merge(cell_range):
        start_cell, end_cell = cell_range.split(':')
        start_col = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = ''.join(filter(str.isalpha, end_cell))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        
        from openpyxl.utils import column_index_from_string
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col_idx, end_col_idx + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        ws.merge_cells(cell_range)
    
    # Título (Fila 1)
    ws['A1'] = 'REGISTRO DE INGRESOS Y EGRESOS DE REACTIVOS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align
    apply_borders_and_merge('A1:H1')
    
    # Subtítulo (Fila 2)
    ws['A2'] = 'CLORO LIBRE / 21055-28'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center_align
    apply_borders_and_merge('A2:H2')
    
    # Encabezados (Fila 3)
    headers = ['Fecha Mes', 'Documento\nSoporte', 'PROVEEDOR /\nSOLICITANTE', 'COD.', 
               'ESPECIFICACION', 'ENTRA\nCant.', 'SALE\nCant.', 'SALDO\nCant.']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = header_fill
        
        # Aplicar colores específicos a las columnas
        if col_idx == 6:  # ENTRA
            cell.fill = entra_fill
        elif col_idx == 7:  # SALE
            cell.fill = sale_fill
        elif col_idx == 8:  # SALDO
            cell.fill = saldo_fill
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Obtener registros del mes
    try:
        fecha_obj = datetime.strptime(fecha_mes, '%Y-%m').date()
        año = fecha_obj.year
        mes = fecha_obj.month
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de fecha inválido. Use YYYY-MM"
        )
    
    # Buscar registros por año y mes
    from sqlalchemy import extract
    registros = db.query(ControlCloroLibre).filter(
        extract('year', ControlCloroLibre.fecha_mes) == año,
        extract('month', ControlCloroLibre.fecha_mes) == mes
    ).order_by(ControlCloroLibre.id).all()
    
    if not registros:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No hay registros para el mes {fecha_mes}"
        )
    
    # Llenar datos
    current_row = 4
    total_entrada = 0
    total_salida = 0
    
    # Primera fila: Inventario inicial (si existe un registro con documento "Inventario")
    inventario_inicial = next((r for r in registros if r.documento_soporte and 'inventario' in r.documento_soporte.lower()), None)
    
    if inventario_inicial:
        entrada_inicial = int(inventario_inicial.cantidad_entra) if inventario_inicial.cantidad_entra else 0
        total_entrada += entrada_inicial  # Incluir en total
        
        ws.cell(row=current_row, column=1).value = ''  # Sin fecha
        ws.cell(row=current_row, column=2).value = inventario_inicial.documento_soporte or 'Inventario'
        ws.cell(row=current_row, column=3).value = inventario_inicial.proveedor_solicitante or ''
        ws.cell(row=current_row, column=4).value = inventario_inicial.codigo or ''
        ws.cell(row=current_row, column=5).value = inventario_inicial.especificacion or ''
        ws.cell(row=current_row, column=6).value = entrada_inicial
        ws.cell(row=current_row, column=7).value = ''
        ws.cell(row=current_row, column=8).value = ''
        
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
        
        # Segunda fila: Solo saldo inicial
        ws.cell(row=current_row, column=8).value = entrada_inicial
        
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
    
    # Filas de datos diarios (excluir el inventario inicial)
    for registro in registros:
        if registro.documento_soporte and 'inventario' in registro.documento_soporte.lower():
            continue  # Saltar el inventario inicial ya procesado
        
        entrada = int(registro.cantidad_entra) if registro.cantidad_entra else 0
        salida = int(registro.cantidad_sale) if registro.cantidad_sale else 0
        saldo = int(registro.cantidad_saldo) if registro.cantidad_saldo else 0
        
        total_entrada += entrada
        total_salida += salida
        
        # Formatear fecha (asumir que cada registro tiene su propia fecha en fecha_mes)
        # En realidad, necesitaríamos un campo fecha_documento para cada día
        ws.cell(row=current_row, column=1).value = registro.fecha_mes.strftime('%d-%b-%y') if registro.fecha_mes else ''
        ws.cell(row=current_row, column=2).value = registro.documento_soporte or ''
        ws.cell(row=current_row, column=3).value = registro.proveedor_solicitante or ''
        ws.cell(row=current_row, column=4).value = registro.codigo or ''
        ws.cell(row=current_row, column=5).value = registro.especificacion or ''
        ws.cell(row=current_row, column=6).value = entrada if entrada > 0 else ''
        ws.cell(row=current_row, column=7).value = salida if salida > 0 else ''
        ws.cell(row=current_row, column=8).value = saldo
        
        # Aplicar estilos a la fila
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
    
    # Fila de totales
    ws.cell(row=current_row, column=1).value = 'TOTALES'
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = center_align
    ws.cell(row=current_row, column=1).fill = total_fill
    apply_borders_and_merge(f'A{current_row}:E{current_row}')
    
    ws.cell(row=current_row, column=6).value = total_entrada
    ws.cell(row=current_row, column=6).font = bold_font
    ws.cell(row=current_row, column=6).alignment = center_align
    ws.cell(row=current_row, column=6).border = thin_border
    ws.cell(row=current_row, column=6).fill = total_fill
    
    ws.cell(row=current_row, column=7).value = total_salida
    ws.cell(row=current_row, column=7).font = bold_font
    ws.cell(row=current_row, column=7).alignment = center_align
    ws.cell(row=current_row, column=7).border = thin_border
    ws.cell(row=current_row, column=7).fill = total_fill
    
    # Última celda de saldo
    ultima_saldo = float(registros[-1].cantidad_saldo) if registros and registros[-1].cantidad_saldo else 0
    ws.cell(row=current_row, column=8).value = ultima_saldo
    ws.cell(row=current_row, column=8).font = bold_font
    ws.cell(row=current_row, column=8).alignment = center_align
    ws.cell(row=current_row, column=8).border = thin_border
    ws.cell(row=current_row, column=8).fill = total_fill
    
    # Guardar en buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Nombre del archivo
    filename = f"registro_reactivos_{fecha_mes}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

