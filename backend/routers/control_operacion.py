"""
Router para operaciones CRUD de Control de Operación (Matriz 2)
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.database import get_db
from models.control_operacion import ControlOperacion
from schemas.control_operacion import (
    ControlOperacionCreate,
    ControlOperacionUpdate,
    ControlOperacionResponse
)

router = APIRouter()


@router.get("/", response_model=List[ControlOperacionResponse])
def get_controles_operacion(
    skip: int = 0,
    limit: int = 100,
    fecha: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """Obtener lista de controles de operación"""
    query = db.query(ControlOperacion)
    
    if fecha:
        query = query.filter(ControlOperacion.fecha == fecha)
    
    controles = query.order_by(
        ControlOperacion.fecha.desc(),
        ControlOperacion.hora.desc()
    ).offset(skip).limit(limit).all()
    
    return controles


@router.get("/fecha/{fecha_consulta}", response_model=List[ControlOperacionResponse])
def get_controles_por_fecha(fecha_consulta: date, db: Session = Depends(get_db)):
    """Obtener todos los controles de un día específico"""
    controles = db.query(ControlOperacion).filter(
        ControlOperacion.fecha == fecha_consulta
    ).order_by(ControlOperacion.hora).all()
    
    return controles


@router.get("/{control_id}", response_model=ControlOperacionResponse)
def get_control_operacion(control_id: int, db: Session = Depends(get_db)):
    """Obtener un control de operación por ID"""
    control = db.query(ControlOperacion).filter(
        ControlOperacion.id == control_id
    ).first()
    
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Control de operación con ID {control_id} no encontrado"
        )
    
    return control


@router.post("/", response_model=ControlOperacionResponse, status_code=status.HTTP_201_CREATED)
def create_control_operacion(control: ControlOperacionCreate, db: Session = Depends(get_db)):
    """Crear un nuevo registro de control de operación"""
    # Verificar si ya existe un registro para esa fecha y hora
    existing = db.query(ControlOperacion).filter(
        ControlOperacion.fecha == control.fecha,
        ControlOperacion.hora == control.hora
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ya existe un registro para {control.fecha} a las {control.hora}"
        )
    
    db_control = ControlOperacion(**control.model_dump())
    db.add(db_control)
    db.commit()
    db.refresh(db_control)
    return db_control


@router.put("/{control_id}", response_model=ControlOperacionResponse)
def update_control_operacion(
    control_id: int,
    control: ControlOperacionUpdate,
    db: Session = Depends(get_db)
):
    """Actualizar un control de operación existente"""
    db_control = db.query(ControlOperacion).filter(
        ControlOperacion.id == control_id
    ).first()
    
    if not db_control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Control de operación con ID {control_id} no encontrado"
        )
    
    update_data = control.model_dump(exclude_unset=True)
    
    for key, value in update_data.items():
        setattr(db_control, key, value)
    
    db.commit()
    db.refresh(db_control)
    return db_control


@router.delete("/{control_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_control_operacion(control_id: int, db: Session = Depends(get_db)):
    """Eliminar un control de operación"""
    db_control = db.query(ControlOperacion).filter(
        ControlOperacion.id == control_id
    ).first()
    
    if not db_control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Control de operación con ID {control_id} no encontrado"
        )
    
    db.delete(db_control)
    db.commit()
    return None


@router.get("/exportar-excel/fecha/{fecha_consulta}")
def exportar_excel_control_operacion(fecha_consulta: date, db: Session = Depends(get_db)):
    """Exportar control de operación a Excel con formato específico"""
    
    # Obtener datos
    controles = db.query(ControlOperacion).filter(
        ControlOperacion.fecha == fecha_consulta
    ).order_by(ControlOperacion.hora).all()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Operación"
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(bold=True, size=10)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Función auxiliar para aplicar bordes antes de merge
    def apply_borders_and_merge(cell_range):
        """Aplica bordes a todas las celdas antes de combinarlas"""
        start_cell, end_cell = cell_range.split(':')
        start_col = start_cell[0]
        end_col = end_cell[0]
        start_row = int(start_cell[1:])
        end_row = int(end_cell[1:])
        
        for row_num in range(start_row, end_row + 1):
            col_ord = ord(start_col)
            while col_ord <= ord(end_col):
                ws[f'{chr(col_ord)}{row_num}'].border = thin_border
                col_ord += 1
        
        ws.merge_cells(cell_range)
    
    # Título principal
    apply_borders_and_merge('A1:T1')
    ws['A1'] = 'PLANTA TRATAMIENTO DE AGUA "LA ESPERANZA"'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = center_alignment
    
    # Subtítulo
    apply_borders_and_merge('A2:T2')
    ws['A2'] = 'CONTROL DE LA OPERACIÓN DEL SISTEMA DE COAGULACION - FLOCULACION Y DEL SISTEMA DE CLORACIÓN'
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = center_alignment
    
    # Encabezado nivel 1 (fila 3) - COAGULACION-FLOCULACION y CLORACION
    # Primero aplicar bordes a todas las celdas de la fila 3
    for col in range(ord('A'), ord('S') + 1):
        ws[f'{chr(col)}3'].border = thin_border
    
    apply_borders_and_merge('F3:L3')
    ws['F3'] = 'COAGULACION - FLOCULACION'
    ws['F3'].font = header_font
    ws['F3'].alignment = center_alignment
    
    apply_borders_and_merge('P3:S3')
    ws['P3'] = 'CLORACIÓN'
    ws['P3'].font = header_font
    ws['P3'].alignment = center_alignment
    
    # Encabezados principales (fila 4)
    headers_row4 = [
        ('A4:A5', 'HORA'),
        ('B4:C4', 'TURBIEDAD'),
        ('D4:D5', 'COLOR'),
        ('E4:G4', 'Ph'),
        ('H4:K4', 'DOSIS QUÍMICOS'),
        ('L4:N4', 'CLARIFICACIÓN'),
        ('O4:O5', 'PRESION\n(PSI)'),
        ('P4:P5', 'PRE\n(Kg/h)'),
        ('Q4:Q5', 'POS\n(Kg/h)'),
        ('R4:R5', 'PRE+POS\n(Kg/h)'),
        ('S4:S5', 'CLORO\nRESIDUAL\n(mg/l)')
    ]
    
    for cell_range, text in headers_row4:
        apply_borders_and_merge(cell_range)
        cell = ws[cell_range.split(':')[0]]
        cell.value = text
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Subencabezados (fila 5)
    subheaders = [
        ('B5', 'AC\n(FTU)'),
        ('C5', 'AT\n(FTU)'),
        ('E5', 'AC'),
        ('F5', 'SULF'),
        ('G5', 'AT'),
        ('H5', 'SULFATO\n(l/s)'),
        ('I5', 'CAL\n(l/s)'),
        ('J5', 'LOERGE\n(l/s)'),
        ('K5', 'FF'),
        ('L5', 'IS\n(K/Cm³)'),
        ('M5', 'CS\n(K/Cm³)'),
        ('N5', 'FS\n(K/Cm³)')
    ]
    
    for cell, text in subheaders:
        ws[cell] = text
        ws[cell].font = header_font
        ws[cell].alignment = center_alignment
        ws[cell].border = thin_border
    
    # Agregar horas (00:00 a 24:00)
    row = 6
    
    # Crear diccionario de controles por hora para búsqueda más eficiente
    controles_por_hora = {}
    for control in controles:
        if control.hora:
            hora_key = f"{control.hora.hour:02d}:00"
            controles_por_hora[hora_key] = control
    
    for hora in range(25):  # 0 a 24
        hora_str = f"{hora:02d}:00"
        ws[f'A{row}'] = hora_str
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = thin_border
        
        # Buscar datos para esta hora
        control = controles_por_hora.get(hora_str)
        
        if control:
            # Convertir Decimal a float para que Excel lo muestre correctamente
            ws[f'B{row}'] = float(control.turbedad_ac) if control.turbedad_ac else None
            ws[f'C{row}'] = float(control.turbedad_at) if control.turbedad_at else None
            ws[f'D{row}'] = control.color
            ws[f'E{row}'] = float(control.ph_ac) if control.ph_ac else None
            ws[f'F{row}'] = float(control.ph_sulf) if control.ph_sulf else None
            ws[f'G{row}'] = float(control.ph_at) if control.ph_at else None
            ws[f'H{row}'] = float(control.dosis_sulfato) if control.dosis_sulfato else None
            ws[f'I{row}'] = float(control.dosis_cal) if control.dosis_cal else None
            ws[f'J{row}'] = float(control.dosis_floergel) if control.dosis_floergel else None
            ws[f'K{row}'] = float(control.ff) if control.ff else None
            ws[f'L{row}'] = float(control.clarificacion_is) if control.clarificacion_is else None
            ws[f'M{row}'] = float(control.clarificacion_cs) if control.clarificacion_cs else None
            ws[f'N{row}'] = float(control.clarificacion_fs) if control.clarificacion_fs else None
            ws[f'O{row}'] = float(control.presion_psi) if control.presion_psi else None
            ws[f'P{row}'] = float(control.presion_pre) if control.presion_pre else None
            ws[f'Q{row}'] = float(control.presion_pos) if control.presion_pos else None
            ws[f'R{row}'] = float(control.presion_total) if control.presion_total else None
            ws[f'S{row}'] = float(control.cloro_residual) if control.cloro_residual else None
        
        # Aplicar bordes a todas las celdas de datos
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Sección inferior con información adicional
    info_row = row + 1
    
    # Fila FECHA y campos relacionados
    ws[f'A{info_row}'] = f'FECHA: {fecha_consulta.strftime("%d/%m/%Y")}'
    ws[f'A{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'J{info_row}:K{info_row}')
    ws[f'J{info_row}'] = 'CANMBIO CIL.: LI'
    ws[f'J{info_row}'].font = Font(bold=True)
    
    ws[f'L{info_row}'] = 'V.'
    ws[f'L{info_row}'].font = Font(bold=True)
    ws[f'L{info_row}'].border = thin_border
    
    ws[f'M{info_row}'] = 'H.'
    ws[f'M{info_row}'].font = Font(bold=True)
    ws[f'M{info_row}'].border = thin_border
    
    ws[f'N{info_row}'] = 'P.'
    ws[f'N{info_row}'].font = Font(bold=True)
    ws[f'N{info_row}'].border = thin_border
    
    # Fila OPERADORES
    info_row += 1
    apply_borders_and_merge(f'A{info_row}:H{info_row}')
    ws[f'A{info_row}'] = 'OPERADORES:'
    ws[f'A{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'I{info_row}:K{info_row}')
    ws[f'I{info_row}'] = 'INICIO DIA:'
    ws[f'I{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'L{info_row}:N{info_row}')
    
    apply_borders_and_merge(f'O{info_row}:Q{info_row}')
    ws[f'O{info_row}'] = 'CILINDROS LLENOS:'
    ws[f'O{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'R{info_row}:S{info_row}')
    
    # Fila REVISADO POR
    info_row += 1
    apply_borders_and_merge(f'A{info_row}:H{info_row}')
    ws[f'A{info_row}'] = 'REVISADO POR:'
    ws[f'A{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'I{info_row}:K{info_row}')
    ws[f'I{info_row}'] = 'CONSUMO DIA:'
    ws[f'I{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'L{info_row}:N{info_row}')
    
    apply_borders_and_merge(f'O{info_row}:Q{info_row}')
    ws[f'O{info_row}'] = 'CILINDROS VACIOS:'
    ws[f'O{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'R{info_row}:S{info_row}')
    
    # Fila ENTRAN
    info_row += 1
    apply_borders_and_merge(f'A{info_row}:H{info_row}')
    ws[f'A{info_row}'] = 'ENTRAN:'
    ws[f'A{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'I{info_row}:K{info_row}')
    ws[f'I{info_row}'] = 'FIN DIA:'
    ws[f'I{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'L{info_row}:N{info_row}')
    
    apply_borders_and_merge(f'O{info_row}:Q{info_row}')
    ws[f'O{info_row}'] = 'CILINDRO EN CONS.:'
    ws[f'O{info_row}'].font = Font(bold=True)
    
    apply_borders_and_merge(f'R{info_row}:S{info_row}')
    
    # Fila OBSERVACION
    info_row += 1
    apply_borders_and_merge(f'A{info_row}:S{info_row}')
    ws[f'A{info_row}'] = 'OBSERVACION:'
    ws[f'A{info_row}'].font = Font(bold=True)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 10
    for col in ['H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
        ws.column_dimensions[col].width = 11
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Retornar como respuesta
    filename = f"control_operacion_{fecha_consulta.strftime('%Y-%m-%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

