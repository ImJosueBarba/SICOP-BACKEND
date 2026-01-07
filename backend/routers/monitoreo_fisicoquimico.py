"""
Router para operaciones CRUD de Monitoreo Fisicoquímico (Matriz 6)
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
from decimal import Decimal
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.database import get_db
from models.monitoreo_fisicoquimico import MonitoreoFisicoquimico
from schemas.monitoreo_fisicoquimico import (
    MonitoreoFisicoquimicoCreate,
    MonitoreoFisicoquimicoUpdate,
    MonitoreoFisicoquimicoResponse
)

router = APIRouter()


@router.get("/", response_model=List[MonitoreoFisicoquimicoResponse])
def get_monitoreos(
    skip: int = 0,
    limit: int = 100,
    fecha: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """Obtener lista de monitoreos fisicoquímicos"""
    query = db.query(MonitoreoFisicoquimico)
    
    if fecha:
        query = query.filter(MonitoreoFisicoquimico.fecha == fecha)
    
    monitoreos = query.order_by(
        MonitoreoFisicoquimico.fecha.desc(),
        MonitoreoFisicoquimico.muestra_numero
    ).offset(skip).limit(limit).all()
    
    return monitoreos


@router.get("/fecha/{fecha_consulta}", response_model=List[MonitoreoFisicoquimicoResponse])
def get_monitoreos_por_fecha(fecha_consulta: date, db: Session = Depends(get_db)):
    """Obtener todos los monitoreos de un día específico (3 muestras)"""
    monitoreos = db.query(MonitoreoFisicoquimico).filter(
        MonitoreoFisicoquimico.fecha == fecha_consulta
    ).order_by(MonitoreoFisicoquimico.muestra_numero).all()
    
    return monitoreos


@router.get("/{monitoreo_id}", response_model=MonitoreoFisicoquimicoResponse)
def get_monitoreo(monitoreo_id: int, db: Session = Depends(get_db)):
    """Obtener un monitoreo por ID"""
    monitoreo = db.query(MonitoreoFisicoquimico).filter(
        MonitoreoFisicoquimico.id == monitoreo_id
    ).first()
    
    if not monitoreo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Monitoreo con ID {monitoreo_id} no encontrado"
        )
    
    return monitoreo


@router.post("/", response_model=MonitoreoFisicoquimicoResponse, status_code=status.HTTP_201_CREATED)
def create_monitoreo(monitoreo: MonitoreoFisicoquimicoCreate, db: Session = Depends(get_db)):
    """Crear un nuevo registro de monitoreo fisicoquímico"""
    # Verificar si ya existe un registro para esa fecha y número de muestra
    existing = db.query(MonitoreoFisicoquimico).filter(
        MonitoreoFisicoquimico.fecha == monitoreo.fecha,
        MonitoreoFisicoquimico.muestra_numero == monitoreo.muestra_numero
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ya existe la muestra {monitoreo.muestra_numero} para {monitoreo.fecha}"
        )
    
    db_monitoreo = MonitoreoFisicoquimico(**monitoreo.model_dump())
    db.add(db_monitoreo)
    db.commit()
    db.refresh(db_monitoreo)
    return db_monitoreo


@router.put("/{monitoreo_id}", response_model=MonitoreoFisicoquimicoResponse)
def update_monitoreo(
    monitoreo_id: int,
    monitoreo: MonitoreoFisicoquimicoUpdate,
    db: Session = Depends(get_db)
):
    """Actualizar un monitoreo existente"""
    db_monitoreo = db.query(MonitoreoFisicoquimico).filter(
        MonitoreoFisicoquimico.id == monitoreo_id
    ).first()
    
    if not db_monitoreo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Monitoreo con ID {monitoreo_id} no encontrado"
        )
    
    update_data = monitoreo.model_dump(exclude_unset=True)
    
    for key, value in update_data.items():
        setattr(db_monitoreo, key, value)
    
    db.commit()
    db.refresh(db_monitoreo)
    return db_monitoreo


@router.delete("/{monitoreo_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_monitoreo(monitoreo_id: int, db: Session = Depends(get_db)):
    """Eliminar un monitoreo"""
    db_monitoreo = db.query(MonitoreoFisicoquimico).filter(
        MonitoreoFisicoquimico.id == monitoreo_id
    ).first()
    
    if not db_monitoreo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Monitoreo con ID {monitoreo_id} no encontrado"
        )
    
    db.delete(db_monitoreo)
    db.commit()
    return None


@router.get("/exportar-excel/fecha/{fecha}")
def exportar_excel(fecha: str, db: Session = Depends(get_db)):
    """Exportar monitoreos fisicoquímicos a Excel con formato específico"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoreo Fisicoquímico"
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Colores
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    agua_cruda_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Azul claro
    agua_tratada_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Verde claro
    
    # Helper function para aplicar bordes y merge
    def apply_borders_and_merge(cell_range):
        start_cell, end_cell = cell_range.split(':')
        start_col = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = ''.join(filter(str.isalpha, end_cell))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        
        from openpyxl.utils import column_index_from_string
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col_idx, end_col_idx + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        ws.merge_cells(cell_range)
    
    # Título (Fila 1)
    ws['A1'] = 'REPORTE DIARIO DE MONITOREO FISICOQUÍMICO COMPLEMENTARIO - AGUA CRUDA/AGUA TRATADA'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = header_fill
    apply_borders_and_merge('A1:M1')
    
    # Fila 2: FECHA
    ws['A2'] = 'FECHA:'
    ws['A2'].font = bold_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A2'].border = thin_border
    apply_borders_and_merge('B2:M2')
    
    # Fila 3: OPERADOR
    ws['A3'] = 'OPERADOR:'
    ws['A3'].font = bold_font
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A3'].border = thin_border
    apply_borders_and_merge('B3:M3')
    
    # Fila 4: LUGAR
    ws['A4'] = 'LUGAR:'
    ws['A4'].font = bold_font
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A4'].border = thin_border
    
    ws['B4'] = 'AGUA CRUDA: ENTRADA A PLANTA LA ESPERANZA'
    ws['B4'].alignment = center_align
    ws['B4'].border = thin_border
    ws['B4'].fill = agua_cruda_fill
    apply_borders_and_merge('B4:F4')
    
    ws['G4'] = 'AGUA TRATADA: TANQUE DE ALMACENAMIENTO PLANTA LA ESPERANZA'
    ws['G4'].alignment = center_align
    ws['G4'].border = thin_border
    ws['G4'].fill = agua_tratada_fill
    apply_borders_and_merge('G4:L4')
    
    ws['M4'].border = thin_border
    
    # Fila 5: Encabezados principales
    ws['A5'] = 'MUESTRA\nNº'
    ws['A5'].font = bold_font
    ws['A5'].alignment = center_align
    ws['A5'].border = thin_border
    ws['A5'].fill = header_fill
    
    ws['B5'] = 'HORA'
    ws['B5'].font = bold_font
    ws['B5'].alignment = center_align
    ws['B5'].border = thin_border
    ws['B5'].fill = header_fill
    
    ws['C5'] = 'AGUA CRUDA'
    ws['C5'].font = bold_font
    ws['C5'].alignment = center_align
    ws['C5'].fill = agua_cruda_fill
    apply_borders_and_merge('C5:G5')
    
    ws['H5'] = 'AGUA TRATADA'
    ws['H5'].font = bold_font
    ws['H5'].alignment = center_align
    ws['H5'].fill = agua_tratada_fill
    apply_borders_and_merge('H5:L5')
    
    ws['M5'] = 'OBSERVACIONES'
    ws['M5'].font = bold_font
    ws['M5'].alignment = center_align
    ws['M5'].border = thin_border
    ws['M5'].fill = header_fill
    
    # Fila 6: Subencabezados
    subheaders = [
        'MUESTRA\nNº', 'HORA',
        'pH', 'CE\n(μs/cm)', 'TDS (ppm)', 'SALIN\n(ppt)', 'TEMP °C',
        'pH', 'CE\n(μs/cm)', 'TDS (ppm)', 'SALIN\n(ppt)', 'TEMP °C',
        'OBSERVACIONES'
    ]
    
    for col_idx, header in enumerate(subheaders, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
        
        # Colores específicos
        if 3 <= col_idx <= 7:  # Agua Cruda
            cell.fill = agua_cruda_fill
        elif 8 <= col_idx <= 12:  # Agua Tratada
            cell.fill = agua_tratada_fill
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 30
    
    # Alto de filas para encabezados
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 40
    
    # Obtener registros de la fecha
    try:
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de fecha inválido. Use YYYY-MM-DD"
        )
    
    registros = db.query(MonitoreoFisicoquimico).filter(
        MonitoreoFisicoquimico.fecha == fecha_obj
    ).order_by(MonitoreoFisicoquimico.muestra_numero).all()
    
    if not registros:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No hay registros para la fecha {fecha}"
        )
    
    # Obtener nombre del operador del primer registro
    operador_nombre = "N/A"
    if registros[0].usuario:
        operador_nombre = registros[0].usuario.nombre_completo
    
    # Llenar FECHA y OPERADOR
    ws['B2'] = fecha_obj.strftime('%d-%m-%Y')
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B3'] = operador_nombre
    ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Llenar datos (3 muestras: 9:00, 12:00, 18:00)
    current_row = 7
    
    for registro in registros:
        # Convertir Decimal a float
        ac_ph = float(registro.ac_ph) if registro.ac_ph else ''
        ac_ce = float(registro.ac_ce) if registro.ac_ce else ''
        ac_tds = float(registro.ac_tds) if registro.ac_tds else ''
        ac_salin = float(registro.ac_salinidad) if registro.ac_salinidad else ''
        ac_temp = float(registro.ac_temperatura) if registro.ac_temperatura else ''
        
        at_ph = float(registro.at_ph) if registro.at_ph else ''
        at_ce = float(registro.at_ce) if registro.at_ce else ''
        at_tds = float(registro.at_tds) if registro.at_tds else ''
        at_salin = float(registro.at_salinidad) if registro.at_salinidad else ''
        at_temp = float(registro.at_temperatura) if registro.at_temperatura else ''
        
        # Llenar fila
        ws.cell(row=current_row, column=1).value = registro.muestra_numero
        ws.cell(row=current_row, column=2).value = registro.hora.strftime('%H:%M:%S') if registro.hora else ''
        ws.cell(row=current_row, column=3).value = ac_ph
        ws.cell(row=current_row, column=4).value = ac_ce
        ws.cell(row=current_row, column=5).value = ac_tds
        ws.cell(row=current_row, column=6).value = ac_salin
        ws.cell(row=current_row, column=7).value = ac_temp
        ws.cell(row=current_row, column=8).value = at_ph
        ws.cell(row=current_row, column=9).value = at_ce
        ws.cell(row=current_row, column=10).value = at_tds
        ws.cell(row=current_row, column=11).value = at_salin
        ws.cell(row=current_row, column=12).value = at_temp
        ws.cell(row=current_row, column=13).value = registro.observaciones or ''
        
        # Aplicar estilos a la fila
        for col in range(1, 14):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
    
    # Pie de página: OPERADOR RESPONSABLE
    footer_row = current_row + 2
    ws.cell(row=footer_row, column=1).value = 'OPERADOR RESPONSABLE'
    ws.cell(row=footer_row, column=1).font = bold_font
    ws.cell(row=footer_row, column=1).alignment = center_align
    apply_borders_and_merge(f'A{footer_row}:M{footer_row}')
    
    # Guardar en buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Nombre del archivo
    filename = f"monitoreo_fisicoquimico_{fecha}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

