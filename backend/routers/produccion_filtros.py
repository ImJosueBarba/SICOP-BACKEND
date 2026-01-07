"""
Router para operaciones CRUD de Producción de Filtros (Matriz 3)
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
from decimal import Decimal
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.database import get_db
from models.produccion_filtro import ProduccionFiltro
from schemas.produccion_filtro import (
    ProduccionFiltroCreate,
    ProduccionFiltroUpdate,
    ProduccionFiltroResponse
)

router = APIRouter()


@router.get("/", response_model=List[ProduccionFiltroResponse])
def get_producciones(
    skip: int = 0,
    limit: int = 100,
    fecha: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """Obtener lista de producciones por filtro"""
    query = db.query(ProduccionFiltro)
    
    if fecha:
        query = query.filter(ProduccionFiltro.fecha == fecha)
    
    producciones = query.order_by(
        ProduccionFiltro.fecha.desc(),
        ProduccionFiltro.hora.desc()
    ).offset(skip).limit(limit).all()
    
    return producciones


@router.get("/fecha/{fecha_consulta}", response_model=List[ProduccionFiltroResponse])
def get_producciones_por_fecha(fecha_consulta: date, db: Session = Depends(get_db)):
    """Obtener todas las producciones de un día específico"""
    producciones = db.query(ProduccionFiltro).filter(
        ProduccionFiltro.fecha == fecha_consulta
    ).order_by(ProduccionFiltro.hora).all()
    
    return producciones


@router.get("/{produccion_id}", response_model=ProduccionFiltroResponse)
def get_produccion(produccion_id: int, db: Session = Depends(get_db)):
    """Obtener una producción por ID"""
    produccion = db.query(ProduccionFiltro).filter(
        ProduccionFiltro.id == produccion_id
    ).first()
    
    if not produccion:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Producción con ID {produccion_id} no encontrada"
        )
    
    return produccion


@router.post("/", response_model=ProduccionFiltroResponse, status_code=status.HTTP_201_CREATED)
def create_produccion(produccion: ProduccionFiltroCreate, db: Session = Depends(get_db)):
    """Crear un nuevo registro de producción"""
    # Verificar si ya existe un registro para esa fecha y hora
    existing = db.query(ProduccionFiltro).filter(
        ProduccionFiltro.fecha == produccion.fecha,
        ProduccionFiltro.hora == produccion.hora
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ya existe un registro para {produccion.fecha} a las {produccion.hora}"
        )
    
    db_produccion = ProduccionFiltro(**produccion.model_dump())
    db.add(db_produccion)
    db.commit()
    db.refresh(db_produccion)
    return db_produccion


@router.put("/{produccion_id}", response_model=ProduccionFiltroResponse)
def update_produccion(
    produccion_id: int,
    produccion: ProduccionFiltroUpdate,
    db: Session = Depends(get_db)
):
    """Actualizar una producción existente"""
    db_produccion = db.query(ProduccionFiltro).filter(
        ProduccionFiltro.id == produccion_id
    ).first()
    
    if not db_produccion:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Producción con ID {produccion_id} no encontrada"
        )
    
    update_data = produccion.model_dump(exclude_unset=True)
    
    for key, value in update_data.items():
        setattr(db_produccion, key, value)
    
    db.commit()
    db.refresh(db_produccion)
    return db_produccion


@router.delete("/{produccion_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_produccion(produccion_id: int, db: Session = Depends(get_db)):
    """Eliminar una producción"""
    db_produccion = db.query(ProduccionFiltro).filter(
        ProduccionFiltro.id == produccion_id
    ).first()
    
    if not db_produccion:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Producción con ID {produccion_id} no encontrada"
        )
    
    db.delete(db_produccion)
    db.commit()
    return None


@router.get("/exportar-excel/fecha/{fecha}")
def exportar_excel(fecha: str, db: Session = Depends(get_db)):
    """Exportar producción por filtros a Excel con formato específico"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Producción Filtros"
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Colores
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Helper function para aplicar bordes y merge
    def apply_borders_and_merge(cell_range):
        start_cell, end_cell = cell_range.split(':')
        start_col = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = ''.join(filter(str.isalpha, end_cell))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        
        from openpyxl.utils import column_index_from_string
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col_idx, end_col_idx + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        ws.merge_cells(cell_range)
    
    # Título (Fila 1)
    ws['A1'] = 'MANCOMUNIDAD LA ESPERANZA'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    apply_borders_and_merge('A1:P1')
    
    # Subtítulo 1 (Fila 2)
    ws['A2'] = 'CONTROL DE PRODUCCIÓN POR FILTRO'
    ws['A2'].font = title_font
    ws['A2'].alignment = center_align
    apply_borders_and_merge('A2:P2')
    
    # Subtítulo 2 (Fila 3)
    ws['A3'] = 'PLANTA DE TRATAMIENTO LA ESPERANZA'
    ws['A3'].font = bold_font
    ws['A3'].alignment = center_align
    apply_borders_and_merge('A3:P3')
    
    # Fila 4: FECHA y OPERADORES
    ws['A4'] = 'Fecha:'
    ws['A4'].font = bold_font
    ws['A4'].border = thin_border
    ws['B4'].border = thin_border
    ws['B4'].fill = yellow_fill
    
    ws['C4'] = 'Operadores:'
    ws['C4'].font = bold_font
    ws['C4'].border = thin_border
    apply_borders_and_merge('D4:P4')
    ws['D4'].border = thin_border
    
    # Encabezados (Fila 5)
    # Columna A: HORA (merge vertical con filas 6 y 7)
    ws['A5'] = 'HORA'
    ws['A5'].font = bold_font
    ws['A5'].alignment = center_align
    ws['A5'].fill = header_fill
    apply_borders_and_merge('A5:A7')
    
    # Filtros 1-6 (cada uno ocupa 2 columnas: h y q)
    filtro_headers = ['FILTRO 1', 'FILTRO 2', 'FILTRO 3', 'FILTRO 4', 'FILTRO 5', 'FILTRO 6']
    start_cols = [2, 4, 6, 8, 10, 12]  # B, D, F, H, J, L (cada filtro ocupa 2 cols)
    
    for idx, (filtro, start_col) in enumerate(zip(filtro_headers, start_cols)):
        cell = ws.cell(row=5, column=start_col)
        cell.value = filtro
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        apply_borders_and_merge(f'{get_column_letter(start_col)}5:{get_column_letter(start_col+1)}5')
    
    # TOTAL (3 columnas)
    ws.cell(row=5, column=14).value = 'TOTAL'
    ws.cell(row=5, column=14).font = bold_font
    ws.cell(row=5, column=14).alignment = center_align
    ws.cell(row=5, column=14).fill = header_fill
    apply_borders_and_merge(f'N5:P5')
    
    # Subencabezados Fila 6 y 7
    # Para cada filtro: h | q en fila 6, y cm | l/s en fila 7
    for start_col in start_cols:
        # Fila 6: h | q
        ws.cell(row=6, column=start_col).value = 'h'
        ws.cell(row=6, column=start_col).font = bold_font
        ws.cell(row=6, column=start_col).alignment = center_align
        ws.cell(row=6, column=start_col).border = thin_border
        ws.cell(row=6, column=start_col).fill = header_fill
        
        ws.cell(row=6, column=start_col+1).value = 'q'
        ws.cell(row=6, column=start_col+1).font = bold_font
        ws.cell(row=6, column=start_col+1).alignment = center_align
        ws.cell(row=6, column=start_col+1).border = thin_border
        ws.cell(row=6, column=start_col+1).fill = header_fill
        
        # Fila 7: cm | l/s
        ws.cell(row=7, column=start_col).value = 'cm'
        ws.cell(row=7, column=start_col).font = bold_font
        ws.cell(row=7, column=start_col).alignment = center_align
        ws.cell(row=7, column=start_col).border = thin_border
        ws.cell(row=7, column=start_col).fill = header_fill
        
        ws.cell(row=7, column=start_col+1).value = 'l/s'
        ws.cell(row=7, column=start_col+1).font = bold_font
        ws.cell(row=7, column=start_col+1).alignment = center_align
        ws.cell(row=7, column=start_col+1).border = thin_border
        ws.cell(row=7, column=start_col+1).fill = header_fill
    
    # TOTAL headers (filas 6-7, merged verticalmente)
    # h (merge filas 6-7)
    ws.cell(row=6, column=14).value = 'h'
    ws.cell(row=6, column=14).font = bold_font
    ws.cell(row=6, column=14).alignment = center_align
    ws.cell(row=6, column=14).fill = header_fill
    apply_borders_and_merge('N6:N7')
    
    # Q (merge filas 6-7)
    ws.cell(row=6, column=15).value = 'Q'
    ws.cell(row=6, column=15).font = bold_font
    ws.cell(row=6, column=15).alignment = center_align
    ws.cell(row=6, column=15).fill = header_fill
    apply_borders_and_merge('O6:O7')
    
    # l/s (merge filas 6-7)
    ws.cell(row=6, column=16).value = 'l/s'
    ws.cell(row=6, column=16).font = bold_font
    ws.cell(row=6, column=16).alignment = center_align
    ws.cell(row=6, column=16).fill = header_fill
    apply_borders_and_merge('P6:P7')
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    for col in range(4, 17):  # D hasta P
        ws.column_dimensions[get_column_letter(col)].width = 6
    
    # Obtener registros de la fecha
    try:
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de fecha inválido. Use YYYY-MM-DD"
        )
    
    registros = db.query(ProduccionFiltro).filter(
        ProduccionFiltro.fecha == fecha_obj
    ).order_by(ProduccionFiltro.hora).all()
    
    # Obtener operadores (si hay registros)
    operador_nombre = "N/A"
    if registros and registros[0].usuario:
        operador_nombre = registros[0].usuario.nombre_completo
    
    # Llenar FECHA y OPERADORES
    ws['B4'] = fecha_obj.strftime('%d-%m-%Y')
    ws['D4'] = operador_nombre
    
    # Crear diccionario de registros por hora
    registros_dict = {}
    for registro in registros:
        if registro.hora:
            # Obtener solo la hora (0-23) y convertir 0 a 24
            hora_num = registro.hora.hour
            if hora_num == 0:
                hora_num = 24
            registros_dict[hora_num] = registro
    
    # Llenar datos (24 horas completas 01H00 a 24H00)
    current_row = 8  # Ahora inicia en fila 8 porque headers ocupan hasta fila 7
    
    for hora in range(1, 25):  # 1 a 24
        hora_display = f"{hora:02d}H00"
        
        # Hora
        ws.cell(row=current_row, column=1).value = hora_display
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=1).alignment = center_align
        
        # Obtener registro si existe para esta hora
        registro = registros_dict.get(hora)
        
        # Filtro 1 (columnas 2-3: B, C) - h | q
        ws.cell(row=current_row, column=2).value = float(registro.filtro1_h) if registro and registro.filtro1_h else ''
        ws.cell(row=current_row, column=3).value = float(registro.filtro1_q) if registro and registro.filtro1_q else ''
        
        # Filtro 2 (columnas 4-5: D, E) - h | q
        ws.cell(row=current_row, column=4).value = float(registro.filtro2_h) if registro and registro.filtro2_h else ''
        ws.cell(row=current_row, column=5).value = float(registro.filtro2_q) if registro and registro.filtro2_q else ''
        
        # Filtro 3 (columnas 6-7: F, G) - h | q
        ws.cell(row=current_row, column=6).value = float(registro.filtro3_h) if registro and registro.filtro3_h else ''
        ws.cell(row=current_row, column=7).value = float(registro.filtro3_q) if registro and registro.filtro3_q else ''
        
        # Filtro 4 (columnas 8-9: H, I) - h | q
        ws.cell(row=current_row, column=8).value = float(registro.filtro4_h) if registro and registro.filtro4_h else ''
        ws.cell(row=current_row, column=9).value = float(registro.filtro4_q) if registro and registro.filtro4_q else ''
        
        # Filtro 5 (columnas 10-11: J, K) - h | q
        ws.cell(row=current_row, column=10).value = float(registro.filtro5_h) if registro and registro.filtro5_h else ''
        ws.cell(row=current_row, column=11).value = float(registro.filtro5_q) if registro and registro.filtro5_q else ''
        
        # Filtro 6 (columnas 12-13: L, M) - h | q
        ws.cell(row=current_row, column=12).value = float(registro.filtro6_h) if registro and registro.filtro6_h else ''
        ws.cell(row=current_row, column=13).value = float(registro.filtro6_q) if registro and registro.filtro6_q else ''
        
        # Total (columnas 14-16: N, O, P) - h | Q | l/s
        ws.cell(row=current_row, column=14).value = ''  # h
        ws.cell(row=current_row, column=15).value = float(registro.caudal_total) if registro and registro.caudal_total else ''  # Q
        ws.cell(row=current_row, column=16).value = ''  # l/s
        
        # Aplicar bordes y alineación a todas las celdas de datos
        for col in range(2, 17):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).alignment = center_align
        
        current_row += 1
    
    # Fila TOTAL (fila 32)
    ws.cell(row=current_row, column=1).value = 'TOTAL'
    ws.cell(row=current_row, column=1).font = bold_font
    ws.cell(row=current_row, column=1).alignment = center_align
    ws.cell(row=current_row, column=1).fill = yellow_fill
    
    # Aplicar bordes y amarillo a toda la fila TOTAL
    for col in range(1, 17):
        ws.cell(row=current_row, column=col).border = thin_border
        ws.cell(row=current_row, column=col).alignment = center_align
        ws.cell(row=current_row, column=col).fill = yellow_fill
    
    # OBSERVACIONES (3 filas combinadas)
    obs_row = current_row + 2
    ws.cell(row=obs_row, column=1).value = 'OBSERVACIONES:'
    ws.cell(row=obs_row, column=1).font = bold_font
    ws.cell(row=obs_row, column=1).alignment = Alignment(horizontal='left', vertical='top')
    
    # Combinar celdas para OBSERVACIONES (3 filas de altura)
    apply_borders_and_merge(f'A{obs_row}:P{obs_row + 2}')
    ws.row_dimensions[obs_row].height = 20
    ws.row_dimensions[obs_row + 1].height = 20
    ws.row_dimensions[obs_row + 2].height = 20
    
    # REVISADO POR (2 filas más abajo)
    rev_row = obs_row + 4
    ws.cell(row=rev_row, column=1).value = 'REVISADO POR:'
    ws.cell(row=rev_row, column=1).font = bold_font
    ws.cell(row=rev_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    apply_borders_and_merge(f'A{rev_row}:P{rev_row}')
    ws.row_dimensions[rev_row].height = 25
    
    # Guardar en buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Nombre del archivo
    filename = f"produccion_filtros_{fecha}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
